# coding=utf-8
import sys
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PIL import Image
import time
from datetime import datetime
from PyQt5.QtCore import QThread
import cv2
import numpy as np

import os
import random

from openpyxl import load_workbook
from openpyxl import Workbook


Alllog = ""

def log(nLog="", detail=""):
    global Alllog
    Alllog += detail + "\n"
    print("LOG[{}] : {}".format(nLog, detail))


# data 파일이 없을 경우
if os.path.isfile("data.xlsx"):
    log(nLog="000", detail="Data File Exist")

else:
    writeExcel = Workbook()
    dataWriteSheet = writeExcel.active
    dataWriteSheet.title = 'data'
    '''
    현재 영상 캡처 저장 경로, 기본 이미지 저장 경로, 너비, 높이, 카메라 번호
    '''

    dataWriteSheet['B2'] = '현재 영상 캡처 저장 경로'
    dataWriteSheet['C2'] = '기본 이미지 저장 경로'
    dataWriteSheet['D2'] = '앱 너비'
    dataWriteSheet['E2'] = '앱 높이'
    dataWriteSheet['F2'] = '카메라 번호'
    dataWriteSheet['G2'] = '캡처 너비'
    dataWriteSheet['H2'] = '캡처 높이'
    dataWriteSheet['I2'] = 'ADAPTIVE_BLOCKSIZE'
    dataWriteSheet['J2'] = 'ADAPTIVE_C'
    dataWriteSheet['K2'] = 'MEDIAN_SIZE'
    dataWriteSheet['L2'] = 'MORPHO_X'
    dataWriteSheet['M2'] = 'MORPHO_Y'
    dataWriteSheet['N2'] = 'CANNY_MIN'
    dataWriteSheet['O2'] = 'CANNY_MAX'
    dataWriteSheet['P2'] = 'OPENCV'


    dataWriteSheet['B3'] = "D:/test/test1111.jpg"
    dataWriteSheet['C3'] = "D:/test/"
    # 664 532
    dataWriteSheet['D3'] = 1680
    dataWriteSheet['E3'] = 1020
    dataWriteSheet['F3'] = 0
    dataWriteSheet['G3'] = 1280
    dataWriteSheet['H3'] = 960
    dataWriteSheet['I3'] = 11
    dataWriteSheet['J3'] = 7
    dataWriteSheet['K3'] = 1
    dataWriteSheet['L3'] = 0
    dataWriteSheet['M3'] = 1
    dataWriteSheet['N3'] = 35
    dataWriteSheet['O3'] = 140
    dataWriteSheet['P3'] = 4

    writeExcel.save(filename='data.xlsx')
    log(nLog="001", detail="Data File No Exist")

# 해당 script 경로의 data.xlsx 파일을 읽어온다.
readExcel = load_workbook(filename='data.xlsx')
dataSheet = readExcel['data']

QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)

'''
현재 영상 캡처 저장 경로, 기본 이미지 저장 경로, 너비, 높이, 카메라 번호
'''
BRING_IN_IMG_ROUTE = dataSheet['B3'].value

SAVE_ORIGIN_IN_IMG_ROUTE = dataSheet['C3'].value
_WIDTH_ADD = int(dataSheet['D3'].value)
_HEIGHT_ADD = int(dataSheet['E3'].value)
CAMERA_NUM = int(dataSheet['F3'].value)
_CAMERA_WIDTH = int(dataSheet['G3'].value)
_CAMERA_HEIGHT = int(dataSheet['H3'].value)
ADAPTIVE_BLOCKSIZE = int(dataSheet['I3'].value)
ADAPTIVE_C = int(dataSheet['J3'].value)
MEDIAN_SIZE = int(dataSheet['K3'].value)
MORPHO_X = int(dataSheet['L3'].value)
MORPHO_Y = int(dataSheet['M3'].value)
CANNY_MIN = int(dataSheet['N3'].value)
CANNY_MAX = int(dataSheet['O3'].value)
nOpencv = int(dataSheet['P3'].value)


log(nLog="002", detail="Below Stored Data Detail : ")

print("현재 영상 캡처 저장 경로 : " + BRING_IN_IMG_ROUTE)
print("기본 이미지 저장 경로 : " + SAVE_ORIGIN_IN_IMG_ROUTE)
print("앱 너비 : " + str(_WIDTH_ADD))
print("앱 높이 : " + str(_HEIGHT_ADD))
print("카메라 번호 : " + str(CAMERA_NUM))
print("카메라 너비 : " + str(_CAMERA_WIDTH))
print("카메라 높이 : " + str(_CAMERA_HEIGHT))
print("ADAPTIVE_BLOCKSIZE : " + str(ADAPTIVE_BLOCKSIZE))
print("ADAPTIVE_C : " + str(ADAPTIVE_C))
print("MEDIAN_SIZE : " + str(MEDIAN_SIZE))
print("MORPHO_X : " + str(MORPHO_X))
print("MORPHO_Y : " + str(MORPHO_Y))
print("CANNY_MIN : " + str(CANNY_MIN))
print("CANNY_MAX : " + str(CANNY_MAX))
print("nOpencv : " + str(nOpencv))

# 664 532

SAVE_IN_IMG_ROUTE = ""

# 자를 영역
area = []

# 현재 영상 캡처본의 이미지 자르기 TOOL 의 UI
class CWidget(QWidget):

    def __init__(self):
        super().__init__()

        global _WIDTH_ADD, _HEIGHT_ADD

        # 캡처를 하면 SUCCESS 가 뜬다.
        self._SUCCESS = ["SUCCESS"]

        self.isSave = False

        # 전체 폼 박스
        formbox = QVBoxLayout()
        self.setLayout(formbox)

        controlLayout = QHBoxLayout()
        right = QHBoxLayout()
        dataInputLayout = QVBoxLayout()

        resultLayout = QHBoxLayout()



        # 자르는 사각형 틀 그래픽 옵션 설정정
        self.pencolor = QColor(180, 85, 162)
        self.brushcolor = QColor(180, 85, 162, 0)

        # left.addStretch(1)

        self.textLineResult = QLineEdit("")
        resultLayout.addWidget(self.textLineResult)

        self.IsResult = False
        self.textResult = QLabel("test1")
        self.textResult.setStyleSheet("color: violet;"
                              "border-style: solid;"
                              "border-width: 2px;"
                              "border-color: violet;"
                              "border-radius: 3px")
        self.textResult.setFont(QFont("맑은 고딕", 12))
        resultLayout.addWidget(self.textResult)

        self.btnSave = QPushButton("Save")
        resultLayout.addWidget(self.btnSave)

        self.btnRetry = QPushButton("Retry")
        resultLayout.addWidget(self.btnRetry)

        resultLayout.setStretchFactor(self.textLineResult, 4)
        resultLayout.setStretchFactor(self.textResult, 1)
        resultLayout.setStretchFactor(self.btnSave, 1)
        resultLayout.setStretchFactor(self.btnRetry, 1)

        # ## control Layout form
        # img Layout
        self.view = CView(self)
        right.addWidget(self.view)

        # data input layout

        self.textOpencvSort = QLabel("Opencv Sort")
        dataInputLayout.addWidget(self.textOpencvSort)

        self.inputOpencvSort = QLineEdit(str(nOpencv))
        dataInputLayout.addWidget(self.inputOpencvSort)

        self.textCannyMin = QLabel("Canny Min")
        dataInputLayout.addWidget(self.textCannyMin)

        self.inputCannyMin = QLineEdit(str(CANNY_MIN))
        dataInputLayout.addWidget(self.inputCannyMin)

        self.textCannyMax = QLabel("Canny Max")
        dataInputLayout.addWidget(self.textCannyMax)

        self.inputCannyMax = QLineEdit(str(CANNY_MAX))
        dataInputLayout.addWidget(self.inputCannyMax)

        self.blank = QLabel("")
        dataInputLayout.addWidget(self.blank)

        dataInputLayout.setStretchFactor(self.blank, 2)

        # 레이아웃 추가
        controlLayout.addLayout(right)
        controlLayout.addLayout(dataInputLayout)

        formbox.addLayout(controlLayout)
        formbox.addLayout(resultLayout)

        # 레이아웃을 formbox 배치
        formbox.setStretchFactor(right, 6)
        formbox.setStretchFactor(resultLayout, 1)

        # 전체 창 크기 설정

        self.setGeometry(0, 0, _WIDTH_ADD, _HEIGHT_ADD)

        self.btnSave.pressed.connect(self.SaveFunction)
        self.btnRetry.pressed.connect(self.RetryFunction)

    def RetryFunction(self):
        global CANNY_MIN, CANNY_MAX, nOpencv

        self.textResult.setText("!!! Retry !!!")

        # readExcel = load_workbook(filename='data.xlsx')
        # dataSheet = readExcel['data']
        global readExcel, dataSheet

        try:
            nOpencv = int(self.inputOpencvSort.text())
            dataSheet['P3'] = nOpencv
        except :
            log(nLog="003", detail="nOpencv Error")

            nOpencv = 0


        try:
            CANNY_MIN = int(self.inputCannyMin.text())
            dataSheet['N3'] = CANNY_MIN

        except:
            log(nLog="004", detail="Canny Min Error")
            CANNY_MIN = 45

        try:
            CANNY_MAX = int(self.inputCannyMax.text())
            dataSheet['O3'] = CANNY_MAX
        except:
            log(nLog="005", detail="Canny Max Error")
            CANNY_MAX = 140

        readExcel.save(filename='./data.xlsx')

        self.view.Capture()
        self.view.ImageRetry()

        log(nLog="006", detail="Press Retry Button")

    def SaveFunction(self):
        global area, SAVE_IN_IMG_ROUTE, SAVE_ORIGIN_IN_IMG_ROUTE
        global CANNY_MIN, CANNY_MAX, nOpencv

        log(nLog="007", detail="Press Save Button")

        _TEST = False

        SaveImg = Image.open(BRING_IN_IMG_ROUTE)

        if not area:
            log(nLog="008", detail="No Set Area")
        else:
            global readExcel, dataSheet

            try:
                nOpencv = int(self.inputOpencvSort.text())
                dataSheet['P3'] = nOpencv
            except:
                log(nLog="009", detail="nOpencv Error")

                nOpencv = 0

            try:
                CANNY_MIN = int(self.inputCannyMin.text())
                dataSheet['N3'] = CANNY_MIN

            except:
                log(nLog="010", detail="Canny Min Error")
                CANNY_MIN = 45

            try:
                CANNY_MAX = int(self.inputCannyMax.text())
                dataSheet['O3'] = CANNY_MAX
            except:
                log(nLog="011", detail="Canny Max Error")
                CANNY_MAX = 140

            readExcel.save(filename='./data.xlsx')

            self.isSave = True

            SAVE_IN_IMG_ROUTE = SAVE_ORIGIN_IN_IMG_ROUTE + datetime.today().strftime("%Y%m%d%H%M%S") + "_{};{};{}".format(nOpencv, CANNY_MIN, CANNY_MAX) + ".jpg"

            try:
                self.textLineResult.setText(SAVE_IN_IMG_ROUTE)

                time.sleep(0.5)
            except:
                print("######################## error1 ######################")

            try:
                self.textResult.setText(self._SUCCESS[random.randint(0, len(self._SUCCESS) - 1)])

            except:
                print("######################## error2 ######################")

            try:
                # area  -  0 : start x / 1 : start y  //  0 : end x / 1 : end y
                crop_area = area
                if crop_area[1] >= crop_area[3] and crop_area[0] <= crop_area[2]:
                    cmp = crop_area[1]
                    crop_area[1] = crop_area[3]
                    crop_area[3] = cmp

                elif crop_area[2] < crop_area[0] and crop_area[3] < crop_area[1]:
                    cmp = crop_area[1]
                    crop_area[1] = crop_area[3]
                    crop_area[3] = cmp
                    cmp = crop_area[0]
                    crop_area[0] = crop_area[2]
                    crop_area[2] = cmp

                elif crop_area[2] < crop_area[0] and crop_area[1] < crop_area[3]:
                    cmp = crop_area[0]
                    crop_area[0] = crop_area[2]
                    crop_area[2] = cmp

                cropped_img = SaveImg.crop(tuple(crop_area))

                cropped_img.save(SAVE_IN_IMG_ROUTE, dpi=(80, 80))

            except:
                print("######################## error3 ######################")



            # print("save")
            # time.sleep(10)
            # sys.exit(app.exec_())


# QGraphicsView display QGraphicsScene
class CView(QGraphicsView, QGraphicsPixmapItem):

    def __init__(self, parent):
        global BRING_IN_IMG_ROUTE

        super().__init__(parent)

        self.Capture()
        # SaveImg = Image.open(BRING_IN_IMG_ROUTE)

        # 그래픽 창 설정
        self.scene = QGraphicsScene()
        #self.img = QGraphicsPixmapItem(QPixmap(frame))
        self.img = QGraphicsPixmapItem(QPixmap(BRING_IN_IMG_ROUTE))

        # 창 그래픽 씬 추가
        self.scene.addItem(self.img)

        self.setScene(self.scene)

        self.items = []

        self.start = QPointF()
        self.end = QPointF()
        self.PenSize = 3

        self.setRenderHint(QPainter.HighQualityAntialiasing)

        self.key = ''

        self._MOVE = 5

        # No Press : 0 / 왼쪽 위 모서리 : 1 / 위 라인 : 2 / 오른쪽 위 모서리 : 3 / 오른쪽 라인 : 4 /
        #
        # 오른쪽 아래 모서리 : 5 / 아래 라인 : 6 / 왼쪽 아래 모서리 : 7 / 왼쪽 라인 : 8
        self.WHERE_PRESSED = 0

    def RemoveDot(self, frame):
        WIDTH = 640
        HEIGHT = 480
        RECT_SIZE = [3, 4, 5]

        OFFSET_WIDTH = []
        OFFSET_HEIGHT = []

        for i in range(0, len(RECT_SIZE), 1):
            OFFSET_WIDTH.append((WIDTH % RECT_SIZE[i]))
            OFFSET_HEIGHT.append((HEIGHT % RECT_SIZE[i]))

        class Position:
            def __init__(self):
                self.x = 0
                self.y = 0

        for i in range(0, len(RECT_SIZE), 1):
            # print('start')
            flag = Position()

            cell = []

            # img pixel reduce
            while flag.y / (HEIGHT - OFFSET_HEIGHT[i]) != 1:

                _cell = []
                flag.x = 0
                while flag.x / (WIDTH - OFFSET_WIDTH[i]) != 1:
                    CHECK = False
                    # 3x3 rect -> 1x1 rect (Reduce)

                    for _x in range(0, RECT_SIZE[i]):
                        if CHECK:
                            break
                        for _y in range(0, RECT_SIZE[i]):
                            try:

                                if frame[flag.y + _y][flag.x + _x] == 0:
                                    CHECK = True
                                    break
                            except:
                                pass

                    # IF CHECK -> image color exist
                    if CHECK:
                        _cell.append(0)
                    else:
                        _cell.append(255)
                    flag.x += RECT_SIZE[i]
                    # END
                cell.append(_cell)
                flag.y += RECT_SIZE[i]

            reviseImg = np.array(cell)
            reviseCompleteImg = reviseImg
            ###

            cur = Position()
            CHECK = True

            # 255 : white , 0 : Black
            # check the isolated dot in the resize img and then erase the isolated dot in the origin img

            while cur.y < int((HEIGHT - OFFSET_HEIGHT[i]) / RECT_SIZE[i]):
                cur.x = 0
                while cur.x < int((WIDTH - OFFSET_WIDTH[i]) / RECT_SIZE[i]):

                    if reviseImg[cur.y][cur.x] == 0:
                        CHECK = True

                        # Find isolated dots
                        for _y in (-1, 0, 1):
                            if CHECK is False:
                                break
                            for _x in (-1, 0, 1):
                                try:
                                    if cur.x + _x >= 0 and cur.y + _y >= 0 and (_x != 0 or _y != 0):
                                        if reviseImg[cur.y + _y][cur.x + _x] < 50:
                                            CHECK = False
                                            break

                                except:
                                    continue

                        # if there is isolated dot, remove it
                        if CHECK:
                            reviseCompleteImg[cur.y][cur.x] = 255

                            for _y in range(0, (RECT_SIZE[i] - 1), 1):
                                for _x in range(0, (RECT_SIZE[i] - 1), 1):
                                    try:
                                        frame[RECT_SIZE[i] * cur.y + _y][RECT_SIZE[i] * cur.x + _x] = 255
                                    except:
                                        pass
                        # Have to Delete area
                        else:

                            CHECK = True

                        if ((HEIGHT - OFFSET_HEIGHT[i]) / int(RECT_SIZE[i])) > (cur.x + 2):
                            cur.x += 2
                            continue

                    cur.x += 1
                cur.y += 1
        return frame

    def Capture(self):

        self.cap = cv2.VideoCapture(CAMERA_NUM)
        self.cap.set(3, _CAMERA_WIDTH)
        self.cap.set(4, _CAMERA_HEIGHT)

        ret, frame = self.cap.read()
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        # frame = cv2.imread('D:/test/test3.jpg', 0)

        log(nLog="012", detail="Image Read({})".format(ret))

        # nOpencv = 5
        global nOpencv, CANNY_MIN, CANNY_MAX

        if nOpencv == 0:
            frame = cv2.adaptiveThreshold(frame, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, ADAPTIVE_BLOCKSIZE, ADAPTIVE_C)
            if MEDIAN_SIZE != 0:
                frame = cv2.medianBlur(frame, MEDIAN_SIZE)
            if MORPHO_X != 0:
                frame = cv2.morphologyEx(frame, cv2.MORPH_CLOSE, np.ones((MORPHO_X, MORPHO_Y), np.uint8))

        elif nOpencv == 1:
            frame = cv2.Canny(frame, CANNY_MIN, CANNY_MAX)
            frame = cv2.adaptiveThreshold(frame, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 9, 3)
        elif nOpencv == 2:
            frame = cv2.GaussianBlur(frame, (1, 1), 15)
            frame = cv2.Canny(frame, CANNY_MIN, CANNY_MAX)
            frame = cv2.adaptiveThreshold(frame, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 13, 15)

        elif nOpencv == 3:

            frame = cv2.Canny(frame, CANNY_MIN, CANNY_MAX)
            #
        elif nOpencv == 4:
            # frame = cv2.GaussianBlur(frame, (1, 1), 3)
            frame = cv2.Canny(frame, CANNY_MIN, CANNY_MAX)
            frame = cv2.morphologyEx(frame, cv2.MORPH_GRADIENT, np.ones((3, 3), np.uint8))
            # frame = cv2.GaussianBlur(frame, (1, 1), 3)
        elif nOpencv == 5:
            frame = cv2.Canny(frame, CANNY_MIN, CANNY_MAX)
            # frame = cv2.adaptiveThreshold(frame, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 9, 5)
            frame = cv2.morphologyEx(frame, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))

        cv2.imwrite(BRING_IN_IMG_ROUTE, frame)

        log(nLog="013", detail="Capture")

    def ImageRetry(self):

        self.img = QGraphicsPixmapItem(QPixmap(BRING_IN_IMG_ROUTE))
        # 창 그래픽 씬 추가
        self.scene.addItem(self.img)
        self.setScene(self.scene)

        self.scene.addItem(self.img)

    def moveEvent(self, e):

        rect = QRectF(self.rect())
        rect.adjust(0, 0, -2, -2)
        # 창 스크롤 바 없애기 위해서 일부 크기 작게 설정
        # rect.setRect(0,0,640,480)

        self.scene.setSceneRect(rect)

    def keyPressEvent(self, e):

        if e.text() == 'a':
            self.key = 'a'
        elif e.text() == 's':
            self.key = 's'
        else:
            self.key = ''

    def keyReleaseEvent(self, e):
        self.key = ''

    def Resize(self):
        global area
        pen = QPen(self.parent().pencolor, self.PenSize)
        brush = QBrush(self.parent().brushcolor)
        pen.setWidth(2)

        self.ImageRetry()

        # if self.start.y() > self.end.y():
        #     cmp = self.end.y()
        #     self.end.setY(self.start.y())
        #     self.start.setY(cmp)

        rect = QRectF(self.start, self.end)

        self.scene.addRect(rect, pen, brush)

        # area = [self.start.x(), self.start.y(), self.end.x(), self.end.y()]
        log(nLog="014", detail="Resize area = {}".format(area))


    def mousePressEvent(self, e):
        #print("aaa")
        global area

        # if self.key == 'a' and e.button() == Qt.LeftButton:
        if self.key == 'a':
            # print("ok")

            point = e.pos()

            # 0 : left / 1 : top / 2 : right / 3 : bottom
            # No Press : 0 / 왼쪽 위 모서리 : 1 / 위 라인 : 2 / 오른쪽 위 모서리 : 3 / 오른쪽 라인 : 4 /
            # 오른쪽 아래 모서리 : 5 / 아래 라인 : 6 / 왼쪽 아래 모서리 : 7 / 왼쪽 라인 : 8

            if (area[0] + self._MOVE > point.x() > area[0] - self._MOVE) and (area[1] + self._MOVE > point.y() > area[1] - self._MOVE):
                # print("왼 쪽 위 모서리를 잡았을 때")
                self.start = QPoint(point.x(), point.y())
                area[0] = point.x()
                area[1] = point.y()
                self.WHERE_PRESSED = 1
                # print("LOG111 : {}".format(area))



            elif (area[2] + self._MOVE > point.x() > area[2] - self._MOVE) and (area[1] + self._MOVE > point.y() > area[1] - self._MOVE):
                # print("오른 쪽 위 모서리를 잡았을 때")
                self.start = QPoint(area[0], point.y())
                self.end = QPoint(point.x(), area[3])

                area[2] = point.x()
                area[1] = point.y()

                self.WHERE_PRESSED = 3

            elif (area[0] + self._MOVE > point.x() > area[0] - self._MOVE) and (area[3] + self._MOVE > point.y() > area[3] - self._MOVE):
                # print("왼 아래 모서리를 잡았을 때")
                self.start = QPoint(point.x(), area[1])
                self.end = QPoint(area[2], point.y())

                area[0] = point.x()
                area[3] = point.y()

                self.WHERE_PRESSED = 7

            elif (area[2] + self._MOVE > point.x() > area[2] - self._MOVE) and (area[3] + self._MOVE > point.y() > area[3] - self._MOVE):
                # print("오른 아래 모서리를 잡았을 때")
                self.end = QPoint(point.x(), point.y())
                area[2] = point.x()
                area[3] = point.y()
                self.WHERE_PRESSED = 5

            else:
                self.WHERE_PRESSED = 0

            self.Resize()

        elif e.button() == Qt.LeftButton and self.key != 'a' and self.key != 's':
            # 시작점 저장
            self.ImageRetry()
            self.start = e.pos()
            self.end = e.pos()
            # print(self.start)

        elif self.key == 's':
            point = e.pos()
            if area[2] > point.x() > area[0] and area[3] > point.y() > area[1]:

                self.WHERE_PRESSED = 8

                center = [int((area[0] + area[2])/2), int((area[1] + area[3])/2)]
                offsetX = point.x() - center[0]
                offsetY = point.y() - center[1]

                self.start = QPoint(area[0] + offsetX, area[1] + offsetY)
                self.end = QPoint(area[2] + offsetX, area[3] + offsetY)
                tmp = [point.x(), point.y()]
                area = [self.start[0], self.start[1], self.end[0], self.end[1]]

                self.Resize()

                log(nLog="120", detail="Move Ret")


    def mouseMoveEvent(self, e):
        # e.buttons()는 정수형 값을 리턴, e.button()은 move시 Qt.Nobutton 리턴
        print("LOG222 : {}".format(self.key))

        if self.key == 'a':
            global area
            # print("ok")

            point = e.pos()
            # print(point)

            # 0 : left / 1 : top / 2 : right / 3 : bottom
            # No Press : 0 / 왼쪽 위 모서리 : 1 / 위 라인 : 2 / 오른쪽 위 모서리 : 3 / 오른쪽 라인 : 4 /
            #
            # 오른쪽 아래 모서리 : 5 / 아래 라인 : 6 / 왼쪽 아래 모서리 : 7 / 왼쪽 라인 : 8

            if self.WHERE_PRESSED == 1:
                # print("왼 쪽 위 모서리를 잡았을 때")
                self.start = QPoint(point.x(), point.y())
                area[0] = point.x()
                area[1] = point.y()

            elif self.WHERE_PRESSED == 3:
                # print("오른 쪽 위 모서리를 잡았을 때")
                self.start = QPoint(area[0], point.y())
                self.end = QPoint(point.x(), area[3])
                area[2] = point.x()
                area[1] = point.y()

            elif self.WHERE_PRESSED == 7:
                # print("왼 아래 모서리를 잡았을 때")
                self.start = QPoint(point.x(), area[1])
                self.end = QPoint(area[2], point.y())
                area[0] = point.x()
                area[3] = point.y()

            elif self.WHERE_PRESSED == 5:
                # print("오른 아래 모서리를 잡았을 때")
                self.end = QPoint(point.x(), point.y())
                area[2] = point.x()
                area[3] = point.y()

            self.Resize()

        elif e.buttons() & Qt.LeftButton and self.key != 'a' and self.key != 's':

            self.end = e.pos()
            pen = QPen(self.parent().pencolor, self.PenSize)
            brush = QBrush(self.parent().brushcolor)
            pen.setWidth(2)

            self.ImageRetry()
            # print("a")
            rect = QRectF(self.start, self.end)
            self.scene.addRect(rect, pen, brush)

        elif self.key == 's':

            point = e.pos()
            if self.WHERE_PRESSED == 8:
                center = [int((area[0] + area[2]) / 2), int((area[1] + area[3]) / 2)]
                offsetX = point.x() - center[0]
                offsetY = point.y() - center[1]

                self.start = QPoint(area[0] + offsetX, area[1] + offsetY)
                self.end = QPoint(area[2] + offsetX, area[3] + offsetY)

                area = [area[0] + offsetX, area[1] + offsetY, area[2] + offsetX, area[3] + offsetY]

                self.Resize()

                log(nLog="120", detail="Move Ret")

    def mouseReleaseEvent(self, e):
        if e.button() == Qt.LeftButton and self.key != 'a':
            global area

            pen = QPen(self.parent().pencolor, self.PenSize)
            brush = QBrush(self.parent().brushcolor)

            self.items.clear()
            rect = QRectF(self.start, self.end)
            self.scene.addRect(rect, pen, brush)

            log(nLog="015", detail="Release : start-[{}, {}] / end-[{}, {}]".format(self.start.x(), self.start.y(), self.end.x(), self.end.y()))

            area = [self.start.x(), self.start.y(), self.end.x(), self.end.y()]


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = CWidget()
    w.show()
    try:
        sys.exit(app.exec_())
    except:
        f = open("log.txt", 'w')
        f.write(Alllog)

        f.close()


